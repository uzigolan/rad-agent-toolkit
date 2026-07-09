"""Build a skill-eval dataset from the ETX-2i show-commands test spreadsheet.

Source: <repo-root>/ETX-2i_Show_Commands_enhanced.xlsx, sheet "STR" — a prior
manual AI-testing log (natural-language prompt -> expected CLI command,
tested across 4 rounds: baseline, post-training, optimized-prompts, retest).

This script filters to rows where Test3 ("Optimized Prompts") == FAIL — cases
that were still getting the wrong CLI syntax even with tuned prompts.

Each "Suggested AI Prompt" cell holds MULTIPLE phrasings, distinguished by
rich-text formatting, not just line breaks:
  - the BOLD run(s) = the "classic" prompt — a direct, explicit phrasing
  - the non-bold run(s) = "implicit" prompts — colloquial phrasings of the
    same intent, harder to match (this is what Test3 "Optimized Prompts"
    was specifically probing). A single non-bold run can bundle several
    newline-separated implicit phrasings.
This distinction is read from the cell's rich-text runs (openpyxl
rich_text=True) — falls back to "first line = classic, rest = implicit" if a
cell ever has no rich-text runs (plain string), so future sheet edits that
lose formatting still degrade gracefully instead of crashing.

Each row's phrasings are expanded into one eval case per phrasing (tagged
classic/implicit), paired with the expected CLI answer from the "Full CLI
Command" column and wrapped as an Abayev/etx2i prompt (triggers the skill's
standing "commands only, never executed" override for that device).

Usage: python scripts/build_eval_dataset.py
Output: tests/eval-etx2i-str-test3-fail.json (+ .csv for human review)
"""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock

REPO = Path(__file__).resolve().parents[1]          # rad-mcp-server/
WORKSPACE = REPO.parent                              # fusion-cli/
XLSX = WORKSPACE / "ETX-2i_Show_Commands_enhanced.xlsx"
OUT_DIR = REPO / "tests"


def split_prompts(cell_value, base_bold: bool) -> tuple[list[str], list[str]]:
    """Split a prompt cell into (classic, implicit) phrasing lists using
    rich-text bold runs. Falls back to first-line/rest if not rich text.

    Newlines are treated as universal line separators regardless of which
    run carries them: some cells have a stray bold-flagged '\\n' run (a
    formatting slip — Enter pressed while still in bold mode) sitting
    between two non-bold lines. Classifying by run-bold alone would file
    that bare newline into the bold bucket and concatenate the two
    surrounding non-bold lines with no separator. Instead, reconstruct
    lines first (newlines always split, whoever carries them), then classify
    each line by the bold-state of its own non-whitespace content."""
    if not isinstance(cell_value, CellRichText):
        lines = [l.strip() for l in (cell_value or "").splitlines() if l.strip()]
        return (lines[:1], lines[1:])

    segments: list[tuple[str, bool | None]] = []  # ("text"|"nl", bold_or_None)
    for run in cell_value:
        if isinstance(run, TextBlock):
            text, bold = run.text, bool(run.font.b)
        else:
            text, bold = str(run), base_bold
        parts = text.split("\n")
        for j, part in enumerate(parts):
            if part != "":
                segments.append((part, bold))
            if j != len(parts) - 1:
                segments.append(("\n", None))

    lines: list[tuple[bool | None, str]] = []
    buf: list[str] = []
    line_bold: bool | None = None
    for text, bold in segments:
        if bold is None:  # newline marker
            lines.append((line_bold, "".join(buf)))
            buf, line_bold = [], None
        else:
            buf.append(text)
            if text.strip():
                line_bold = bold
    lines.append((line_bold, "".join(buf)))

    classic = [t.strip() for b, t in lines if b is True and t.strip()]
    implicit = [t.strip() for b, t in lines if b is False and t.strip()]
    return classic, implicit


def main() -> None:
    wb_rich = openpyxl.load_workbook(XLSX, rich_text=True)
    ws_rich = wb_rich["STR"]
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["STR"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    cases = []
    for r_idx, r in enumerate(rows, start=2):
        (num, category, cli_path, show_cmd, full_cli, prompts_cell,
         t1, t2, t3, t4, comments, *_rest) = r
        if t3 != "FAIL":
            continue

        prompt_cell = ws_rich.cell(row=r_idx, column=6)
        classic, implicit = split_prompts(prompt_cell.value, bool(prompt_cell.font.b))

        tagged = [(p, "classic", i) for i, p in enumerate(classic, 1)] + \
                 [(p, "implicit", i) for i, p in enumerate(implicit, 1)]
        for prompt, ptype, i in tagged:
            cases.append({
                "id": f"str-{num}-{ptype}-{i}",
                "source_row": num,
                "category": category,
                "cli_path": cli_path,
                "expected_cli_command": full_cli,
                "show_command": show_cmd,
                "prompt": prompt,
                "prompt_type": ptype,
                "wrapped_prompt": f"abayev, {prompt} on etx2i",
                "device": "etx2i",
                "prior_test_results": {
                    "test1_baseline": t1, "test2_post_training": t2,
                    "test3_optimized_prompts": t3, "test4_post_training_retest": t4,
                },
                "prior_comments": (comments or "").strip() or None,
            })

    OUT_DIR.mkdir(exist_ok=True)
    json_path = OUT_DIR / "eval-etx2i-str-test3-fail.json"
    json_path.write_text(json.dumps(cases, indent=2), encoding="utf-8")

    csv_path = OUT_DIR / "eval-etx2i-str-test3-fail.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["id", "source_row", "category", "cli_path", "prompt_type",
                    "prompt", "wrapped_prompt", "expected_cli_command", "prior_comments"])
        for c in cases:
            w.writerow([c["id"], c["source_row"], c["category"], c["cli_path"],
                        c["prompt_type"], c["prompt"], c["wrapped_prompt"],
                        c["expected_cli_command"], c["prior_comments"] or ""])

    rows_used = len(set(c["source_row"] for c in cases))
    n_classic = sum(1 for c in cases if c["prompt_type"] == "classic")
    n_implicit = sum(1 for c in cases if c["prompt_type"] == "implicit")
    print(f"{len(cases)} eval cases from {rows_used} Test3-FAIL rows "
          f"({n_classic} classic, {n_implicit} implicit)")
    print(f"-> {json_path.relative_to(REPO)}")
    print(f"-> {csv_path.relative_to(REPO)}")


if __name__ == "__main__":
    main()
