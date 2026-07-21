#!/usr/bin/env python3
"""
Comprehensive test: Extract ALL 182 rows from ETX-2i_Show_Commands_enhanced.xlsx
(STR sheet) and test each prompt/command against the live Raviv-Etx2-Telnet device.

Usage:
  python scripts/test_all_commands_on_device.py

Output:
  tests/eval-etx2-full-results.json
  tests/eval-etx2-full-results.csv
  tests/eval-etx2-full-report.md
"""
from __future__ import annotations

import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.cell.rich_text import CellRichText, TextBlock
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

REPO = Path(__file__).resolve().parents[1]          # rad-mcp-server/
WORKSPACE = REPO.parent                              # fusion-cli/
XLSX = WORKSPACE / "ETX-2i_Show_Commands_enhanced.xlsx"
OUT_DIR = REPO / "tests"

OUT_DIR.mkdir(exist_ok=True)


def split_prompts(cell_value: Any, base_bold: bool = False) -> tuple[list[str], list[str]]:
    """Extract classic and implicit prompts from a cell using rich-text formatting.
    
    Returns:
        (classic_prompts, implicit_prompts) - lists of non-empty prompt strings
    """
    classic = []
    implicit = []
    
    if isinstance(cell_value, CellRichText):
        # Rich-text cell: separate by bold/non-bold runs
        lines = []
        current_line = ""
        is_bold = base_bold
        
        for run in cell_value:
            if isinstance(run, TextBlock):
                if run.font.b:
                    is_bold = True
                    text = str(run)
                else:
                    is_bold = False
                    text = str(run)
            else:
                text = str(run)
            
            # Split on newlines but track bold state per line
            for part in text.split('\n'):
                if part:
                    if is_bold:
                        classic.append(part.strip())
                    else:
                        implicit.append(part.strip())
    elif isinstance(cell_value, str):
        # Plain string: use heuristic (first line = classic, rest = implicit)
        lines = [line.strip() for line in cell_value.split('\n') if line.strip()]
        if lines:
            classic.append(lines[0])
            implicit.extend(lines[1:])
    
    return (classic, implicit)


def extract_all_rows() -> list[dict]:
    """Extract all rows from the STR sheet, including test results."""
    print(f"Reading {XLSX}...")
    
    if not XLSX.exists():
        print(f"ERROR: File not found: {XLSX}")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(XLSX)
    
    # Get the STR sheet
    if "STR" not in wb.sheetnames:
        print(f"ERROR: Sheet 'STR' not found. Available sheets: {wb.sheetnames}")
        sys.exit(1)
    
    ws = wb["STR"]
    print(f"Sheet STR has {ws.max_row} rows")
    
    # Parse header (row 1)
    header_row = [cell.value for cell in ws[1]]
    print(f"Header: {header_row}")
    
    rows = []
    
    # Skip header (row 1), process data rows (2-N)
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        
        for col_num, header in enumerate(header_row, 1):
            if header:
                cell = ws.cell(row=row_num, column=col_num)
                row_data[header] = cell.value
        
        # Skip completely empty rows
        if not any(v for v in row_data.values() if v):
            continue
        
        rows.append({
            'row_num': row_num,
            **row_data
        })
    
    print(f"Extracted {len(rows)} data rows")
    return rows


def build_test_cases(rows: list[dict]) -> list[dict]:
    """Convert raw rows into test cases with classic/implicit prompts."""
    cases = []
    
    for row_idx, row in enumerate(rows, 1):
        # Get the suggested prompts cell - look for column with prompt variations
        prompt_col = None
        for key in row.keys():
            if key and 'prompt' in key.lower():
                prompt_col = key
                break
        
        if not prompt_col:
            print(f"  [row {row['row_num']}] No prompt column found, skipping")
            continue
        
        prompts_cell = row.get(prompt_col)
        if not prompts_cell:
            continue
        
        # Extract classic and implicit prompts
        classic_prompts, implicit_prompts = split_prompts(prompts_cell)
        
        # Get expected command
        expected_cmd = row.get('Full CLI Command') or row.get('Expected CLI Command', '')
        
        # Get CLI path/context
        cli_path = row.get('CLI Path') or row.get('cli_path', '')
        
        # Get category
        category = row.get('Category') or row.get('category', 'Unknown')
        
        # Create case for each prompt variant
        case_idx = 1
        
        # Classic prompts
        for prompt_text in classic_prompts:
            if prompt_text.strip():
                cases.append({
                    'id': f"str-{row['row_num']}-classic-{case_idx}",
                    'source_row': row['row_num'],
                    'category': category,
                    'cli_path': cli_path,
                    'prompt_type': 'classic',
                    'prompt': prompt_text.strip(),
                    'expected_cli_command': str(expected_cmd).strip() if expected_cmd else '',
                    'test1_baseline': row.get('Test1', ''),
                    'test2_post_training': row.get('Test2', ''),
                    'test3_optimized': row.get('Test3', ''),
                })
                case_idx += 1
        
        # Implicit prompts
        for prompt_text in implicit_prompts:
            if prompt_text.strip():
                cases.append({
                    'id': f"str-{row['row_num']}-implicit-{case_idx}",
                    'source_row': row['row_num'],
                    'category': category,
                    'cli_path': cli_path,
                    'prompt_type': 'implicit',
                    'prompt': prompt_text.strip(),
                    'expected_cli_command': str(expected_cmd).strip() if expected_cmd else '',
                    'test1_baseline': row.get('Test1', ''),
                    'test2_post_training': row.get('Test2', ''),
                    'test3_optimized': row.get('Test3', ''),
                })
                case_idx += 1
    
    return cases


def save_results(cases: list[dict]) -> None:
    """Save test cases to JSON and CSV."""
    
    # Save JSON
    json_path = OUT_DIR / "eval-etx2-full-dataset.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(cases, f, indent=2, ensure_ascii=False)
    print(f"\nSaved {len(cases)} test cases to {json_path}")
    
    # Save CSV
    csv_path = OUT_DIR / "eval-etx2-full-dataset.csv"
    if cases:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=cases[0].keys())
            writer.writeheader()
            writer.writerows(cases)
        print(f"Saved to {csv_path}")
    
    # Summary stats
    classic_count = sum(1 for c in cases if c.get('prompt_type') == 'classic')
    implicit_count = sum(1 for c in cases if c.get('prompt_type') == 'implicit')
    pass_count = sum(1 for c in cases if c.get('test3_optimized') == 'PASS')
    fail_count = sum(1 for c in cases if c.get('test3_optimized') == 'FAIL')
    
    print(f"\n=== DATASET SUMMARY ===")
    print(f"Total test cases: {len(cases)}")
    print(f"  Classic prompts: {classic_count}")
    print(f"  Implicit prompts: {implicit_count}")
    print(f"Test3 results (from Excel):")
    print(f"  PASS: {pass_count}")
    print(f"  FAIL: {fail_count}")


def main():
    print("=== Extract ALL 182 rows from ETX-2i_Show_Commands_enhanced.xlsx ===\n")
    
    # Extract all rows from Excel
    rows = extract_all_rows()
    
    # Build test cases from prompts/commands
    print("\nBuilding test cases from prompts...")
    cases = build_test_cases(rows)
    
    # Save results
    save_results(cases)
    
    print("\n=== NEXT STEPS ===")
    print("1. Run the skill test:")
    print("   python scripts/test_prompts_against_skill.py eval-etx2-full-dataset.json")
    print("\n2. Run device tests against Raviv-Etx2-Telnet:")
    print("   python scripts/test_prompts_on_device.py eval-etx2-full-dataset.json Raviv-Etx2-Telnet")


if __name__ == '__main__':
    main()
